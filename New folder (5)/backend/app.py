import os
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import openpyxl
from openpyxl import Workbook
import json
from datetime import datetime

current_dir = os.path.dirname(os.path.abspath(__file__))
frontend_dir = os.path.join(current_dir, '..', 'frontend')

app = Flask(__name__, static_folder=frontend_dir, static_url_path='')
CORS(app)

@app.route('/')
def index():
    return send_from_directory(frontend_dir, 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory(frontend_dir, path)

EXCEL_FILE = "quiz_results.xlsx"
PENDING_FILE = "pending_results.json"

def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["Name", "Score", "Total", "Entry Date", "Entry Time", "Submit Time", "Wrong Answers"])
        wb.save(EXCEL_FILE)

def load_pending():
    if os.path.exists(PENDING_FILE):
        with open(PENDING_FILE, 'r', encoding='utf-8') as f:
            try:
                return json.load(f)
            except json.JSONDecodeError:
                return []
    return []

def save_pending(queue):
    with open(PENDING_FILE, 'w', encoding='utf-8') as f:
        json.dump(queue, f, ensure_ascii=False, indent=2)

@app.route('/api/check-name', methods=['GET'])
def check_name():
    name = request.args.get('name')
    if not name:
        return jsonify({"error": "Missing name parameter"}), 400
    
    # 1. Check pending queue first
    pending = load_pending()
    for item in pending:
        if item[0] and str(item[0]).strip().lower() == str(name).strip().lower():
            return jsonify({"exists": True})
            
    # 2. Check excel
    if not os.path.exists(EXCEL_FILE):
        return jsonify({"exists": False})

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
        ws = wb.active
        
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            if row[0] and str(row[0]).strip().lower() == str(name).strip().lower():
                return jsonify({"exists": True})
                
        return jsonify({"exists": False})
    except Exception as e:
        print("Error reading excel for check-name:", e)
        return jsonify({"exists": False})

@app.route('/api/submit', methods=['POST'])
def submit_score():
    data = request.json
    name = data.get('name')
    score = data.get('score')
    total = data.get('total')
    entry_date = data.get('entryDate', '')
    entry_time = data.get('entryTime', '')
    wrong_answers = data.get('wrongAnswers', '')
    
    if not name or score is None or total is None:
        return jsonify({"error": "Missing fields"}), 400
        
    now = datetime.now()
    submit_time = now.strftime("%H:%M:%S")
    new_row = [name, score, total, entry_date, entry_time, submit_time, wrong_answers]
    
    pending_queue = load_pending()
    pending_queue.append(new_row)
    
    try:
        initialize_excel()
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Flush everything from pending_queue
        for row_data in pending_queue:
            ws.append(row_data)
        
        # Autofit columns manually
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 50
            
        wb.save(EXCEL_FILE)
        
        # If successfully saved, clear the queue
        if os.path.exists(PENDING_FILE):
            os.remove(PENDING_FILE)
            
        return jsonify({"success": True, "message": "Saved successfully!"})
    except PermissionError:
        # File is locked! Save it to pending.json
        save_pending(pending_queue)
        print("Permission error. Excel file is open. Saved to queue!")
        # We return SUCCESS to the frontend so the student doesn't see an error.
        return jsonify({"success": True, "message": "Saved to queue successfully!"})
    except Exception as e:
        print("General Exception in submit:", e)
        return jsonify({"error": str(e)}), 500

@app.route('/api/flush', methods=['GET'])
def flush_queue():
    pending_queue = load_pending()
    if not pending_queue:
        return jsonify({"status": "Queue is empty!"})
        
    try:
        initialize_excel()
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        for row_data in pending_queue:
            ws.append(row_data)
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 50
        wb.save(EXCEL_FILE)
        if os.path.exists(PENDING_FILE):
            os.remove(PENDING_FILE)
        return jsonify({"status": f"Flushed {len(pending_queue)} pending results!"})
    except PermissionError:
        return jsonify({"status": "Failed to flush, Excel file is still open."})
    except Exception as e:
        return jsonify({"status": f"Error: {e}"})

import threading
import time

def background_flusher():
    while True:
        time.sleep(3)
        try:
            pending_queue = load_pending()
            if not pending_queue:
                continue
                
            if not os.path.exists(EXCEL_FILE):
                initialize_excel()
                
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            for row_data in pending_queue:
                ws.append(row_data)
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 50
            wb.save(EXCEL_FILE)
            
            if os.path.exists(PENDING_FILE):
                os.remove(PENDING_FILE)
            print("Background flush successful!")
        except Exception:
            pass

if __name__ == '__main__':
    initialize_excel()
    threading.Thread(target=background_flusher, daemon=True).start()
    print("Backend server is running on http://localhost:5000")
    print("Results will be saved to", os.path.abspath(EXCEL_FILE))
    app.run(port=5000, debug=False)
